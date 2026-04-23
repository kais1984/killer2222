#!/usr/bin/env python3
"""
RIAMAN FASHION ERP System - Excel Workbook Generator
Creates a complete, professionally designed ERP system for luxury fashion retail
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage
import datetime

# Define luxury brand color palette
COLORS = {
    'ivory': 'F5F5F0',
    'beige': 'E8E3D8',
    'champagne': 'F7E9D3',
    'charcoal': '2C2C2C',
    'gold': 'D4A574',
    'light_gray': 'E8E8E8',
    'white': 'FFFFFF',
    'dark_text': '333333'
}

FONTS = {
    'header': Font(name='Calibri', size=14, bold=True, color=COLORS['charcoal']),
    'subheader': Font(name='Calibri', size=11, bold=True, color=COLORS['charcoal']),
    'label': Font(name='Calibri', size=10, color=COLORS['charcoal']),
    'data': Font(name='Calibri', size=10, color=COLORS['dark_text']),
    'gold_header': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
}

def create_luxury_header(sheet, title, subtitle, start_row=1):
    """Create a luxury branded header section"""
    # Merge cells for header
    sheet.merge_cells(f'A{start_row}:H{start_row}')
    header_cell = sheet[f'A{start_row}']
    header_cell.value = 'RIAMAN FASHION'
    header_cell.font = Font(name='Garamond', size=16, bold=True, color=COLORS['gold'])
    header_cell.fill = PatternFill(start_color=COLORS['charcoal'], end_color=COLORS['charcoal'], fill_type='solid')
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[start_row].height = 30

    # Subtitle
    start_row += 1
    sheet.merge_cells(f'A{start_row}:H{start_row}')
    subtitle_cell = sheet[f'A{start_row}']
    subtitle_cell.value = title
    subtitle_cell.font = Font(name='Calibri', size=12, bold=True, color=COLORS['gold'])
    subtitle_cell.fill = PatternFill(start_color=COLORS['ivory'], end_color=COLORS['ivory'], fill_type='solid')
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[start_row].height = 25

    # Description
    if subtitle:
        start_row += 1
        sheet.merge_cells(f'A{start_row}:H{start_row}')
        desc_cell = sheet[f'A{start_row}']
        desc_cell.value = subtitle
        desc_cell.font = Font(name='Calibri', size=9, italic=True, color=COLORS['charcoal'])
        desc_cell.fill = PatternFill(start_color=COLORS['beige'], end_color=COLORS['beige'], fill_type='solid')
        desc_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[start_row].height = 20

    return start_row + 2

def apply_table_style(sheet, start_row, end_row, start_col=1, end_col=8):
    """Apply luxury table styling"""
    # Header row (start_row - 1)
    header_fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    header_font = FONTS['gold_header']
    
    thin_border = Border(
        left=Side(style='thin', color=COLORS['light_gray']),
        right=Side(style='thin', color=COLORS['light_gray']),
        top=Side(style='thin', color=COLORS['light_gray']),
        bottom=Side(style='thin', color=COLORS['light_gray'])
    )

    # Apply alternating row colors
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border
            
            # Alternate row colors
            if (row - start_row) % 2 == 0:
                cell.fill = PatternFill(start_color=COLORS['white'], end_color=COLORS['white'], fill_type='solid')
            else:
                cell.fill = PatternFill(start_color=COLORS['ivory'], end_color=COLORS['ivory'], fill_type='solid')
            
            cell.font = FONTS['data']

def create_workbook():
    """Create the complete RIAMAN FASHION ERP workbook"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    print("Creating RIAMAN FASHION ERP Workbook...")
    
    # ============================================================================
    # BACKEND SHEETS (Hidden)
    # ============================================================================
    
    # 1. _System Sheet
    ws_system = wb.create_sheet("_System")
    ws_system.sheet_state = 'hidden'
    ws_system['A1'].value = "System Configuration"
    ws_system['A2'].value = "App Name"
    ws_system['B2'].value = "RIAMAN FASHION ERP"
    ws_system['A3'].value = "Version"
    ws_system['B3'].value = "1.0"
    ws_system['A4'].value = "Created"
    ws_system['B4'].value = datetime.datetime.now()
    print("✓ _System sheet created")
    
    # 2. _Products_Master Sheet
    ws_products = wb.create_sheet("_Products_Master")
    ws_products.sheet_state = 'hidden'
    headers = ["Product ID", "Name", "Category", "Size", "Color", "Fabric", "Designer/Collection",
               "Cost Price (AED)", "Selling Price (AED)", "Rental Price (AED/day)", "Deposit (AED)",
               "Condition", "Availability", "Created Date", "Image Path", "Barcode"]
    ws_products.append(headers)
    
    # Style header
    for col, header in enumerate(headers, 1):
        cell = ws_products.cell(row=1, column=col)
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    # Add sample product
    ws_products.append([
        "RIAMAN-00001",
        "Eternal Dreams Bridal Gown",
        "Wedding",
        "US 6",
        "White",
        "Silk Tulle",
        "RIAMAN Couture",
        3500,
        8500,
        800,
        2500,
        "New",
        "Available",
        datetime.datetime.now(),
        "[image_path]",
        "RIAMAN-00001"
    ])
    
    # Add date picker for Created Date column (column N)
    prod_date_validation = DataValidation(type="date", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
    prod_date_validation.error = 'Please enter a valid date'
    prod_date_validation.errorTitle = 'Invalid Date'
    ws_products.add_data_validation(prod_date_validation)
    prod_date_validation.add('N2:N1000')
    
    print("✓ _Products_Master sheet created with date picker")
    
    # 3. _Stock_Movements Sheet
    ws_movements = wb.create_sheet("_Stock_Movements")
    ws_movements.sheet_state = 'hidden'
    headers = ["Movement ID", "Product ID", "Movement Type", "Quantity", "Reference", "Notes", "Date", "User"]
    ws_movements.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_movements.cell(row=1, column=col)
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    # Add date picker for Date column (column G)
    mov_backend_validation = DataValidation(type="date", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
    mov_backend_validation.error = 'Please enter a valid date'
    mov_backend_validation.errorTitle = 'Invalid Date'
    ws_movements.add_data_validation(mov_backend_validation)
    mov_backend_validation.add('G2:G1000')
    
    print("✓ _Stock_Movements sheet created with date picker")
    
    # 4. _Stock_Current Sheet
    ws_stock = wb.create_sheet("_Stock_Current")
    ws_stock.sheet_state = 'hidden'
    headers = ["Product ID", "Product Name", "Current Stock", "Low Stock Alert", "Last Updated"]
    ws_stock.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_stock.cell(row=1, column=col)
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    print("✓ _Stock_Current sheet created")
    
    # 5. _Customers_Master Sheet
    ws_customers = wb.create_sheet("_Customers_Master")
    ws_customers.sheet_state = 'hidden'
    headers = ["Customer ID", "Name", "Phone", "Email", "Notes", "Total Spending (AED)", "Visit Count",
               "Preferences", "VIP Status", "Created Date", "Last Visit"]
    ws_customers.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_customers.cell(row=1, column=col)
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    # Add date pickers for Created Date (column J) and Last Visit (column K)
    cust_date_validation = DataValidation(type="date", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
    cust_date_validation.error = 'Please enter a valid date'
    cust_date_validation.errorTitle = 'Invalid Date'
    ws_customers.add_data_validation(cust_date_validation)
    cust_date_validation.add('J2:J1000')
    cust_date_validation.add('K2:K1000')
    
    print("✓ _Customers_Master sheet created with date pickers")
    
    # 6. _Sales_Transactions Sheet
    ws_sales = wb.create_sheet("_Sales_Transactions")
    ws_sales.sheet_state = 'hidden'
    headers = ["Transaction ID", "Customer ID", "Transaction Date", "Payment Method", "Subtotal (AED)",
               "VAT (AED)", "Discount (AED)", "User", "Total (AED)", "Completed Time"]
    ws_sales.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_sales.cell(row=1, column=col)
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    # Add formulas and date picker for Transaction Date (column C)
    sales_date_validation = DataValidation(type="date", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
    sales_date_validation.error = 'Please enter a valid date'
    sales_date_validation.errorTitle = 'Invalid Date'
    ws_sales.add_data_validation(sales_date_validation)
    sales_date_validation.add('C2:C1000')
    
    print("✓ _Sales_Transactions sheet created with date picker")
    
    # 7. _Transaction_Items Sheet
    ws_items = wb.create_sheet("_Transaction_Items")
    ws_items.sheet_state = 'hidden'
    headers = ["Transaction ID", "Product ID", "Quantity", "Unit Price (AED)", "Is Rental", "Rental Days", "Line Total (AED)"]
    ws_items.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_items.cell(row=1, column=col)
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    print("✓ _Transaction_Items sheet created")
    
    # 8. _Rental_Bookings Sheet
    ws_rentals = wb.create_sheet("_Rental_Bookings")
    ws_rentals.sheet_state = 'hidden'
    headers = ["Rental ID", "Product ID", "Customer ID", "Start Date", "End Date", "Deposit (AED)",
               "Status", "Created Date", "Late Fee (AED)", "Return Date", "Condition on Return"]
    ws_rentals.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_rentals.cell(row=1, column=col)
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    # Add date pickers for all date columns: Start Date (D), End Date (E), Created Date (H), Return Date (J)
    rental_date_validation = DataValidation(type="date", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
    rental_date_validation.error = 'Please enter a valid date'
    rental_date_validation.errorTitle = 'Invalid Date'
    ws_rentals.add_data_validation(rental_date_validation)
    rental_date_validation.add('D2:D1000')
    rental_date_validation.add('E2:E1000')
    rental_date_validation.add('H2:H1000')
    rental_date_validation.add('J2:J1000')
    
    print("✓ _Rental_Bookings sheet created with date pickers")
    
    # 9. _Accounting_Ledger Sheet
    ws_accounting = wb.create_sheet("_Accounting_Ledger")
    ws_accounting.sheet_state = 'hidden'
    headers = ["Reference", "Account Type", "Debit (AED)", "Credit (AED)", "Date", "Description", "User"]
    ws_accounting.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_accounting.cell(row=1, column=col)
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    # Add date picker for Date column (column E)
    acct_date_validation = DataValidation(type="date", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
    acct_date_validation.error = 'Please enter a valid date'
    acct_date_validation.errorTitle = 'Invalid Date'
    ws_accounting.add_data_validation(acct_date_validation)
    acct_date_validation.add('E2:E1000')
    
    print("✓ _Accounting_Ledger sheet created with date picker")
    
    # 10. _Audit_Log Sheet
    ws_audit = wb.create_sheet("_Audit_Log")
    ws_audit.sheet_state = 'hidden'
    headers = ["Timestamp", "User", "Action", "Reference", "Details"]
    ws_audit.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_audit.cell(row=1, column=col)
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    print("✓ _Audit_Log sheet created")
    
    # 11. _Users_Master Sheet
    ws_users = wb.create_sheet("_Users_Master")
    ws_users.sheet_state = 'hidden'
    headers = ["User ID", "Username", "Password (Hashed)", "Full Name", "Role", "Email", "Active", "Created Date"]
    ws_users.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws_users.cell(row=1, column=col)
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    # Add default users
    ws_users.append(["USR001", "admin", "admin123", "Administrator", "Admin", "admin@riaman.ae", True, datetime.datetime.now()])
    ws_users.append(["USR002", "sales", "sales123", "Sales Staff", "Sales Person", "sales@riaman.ae", True, datetime.datetime.now()])
    ws_users.append(["USR003", "stock", "stock123", "Stock Manager", "Stock Manager", "stock@riaman.ae", True, datetime.datetime.now()])
    ws_users.append(["USR004", "accounts", "accounts123", "Accountant", "Accountant", "accounts@riaman.ae", True, datetime.datetime.now()])
    
    print("✓ _Users_Master sheet created")
    
    # 12. _Settings Sheet
    ws_settings = wb.create_sheet("_Settings")
    ws_settings.sheet_state = 'hidden'
    ws_settings['A1'].value = "Setting"
    ws_settings['B1'].value = "Value"
    
    settings_data = [
        ["VIP Threshold (AED)", 50000],
        ["VAT Rate (%)", 5],
        ["Default Rental Late Fee (AED/day)", 50],
        ["Low Stock Alert Level", 2],
        ["Company Name", "RIAMAN FASHION"],
        ["Currency", "AED"],
        ["Country", "UAE"]
    ]
    
    for row, data in enumerate(settings_data, 2):
        ws_settings[f'A{row}'].value = data[0]
        ws_settings[f'B{row}'].value = data[1]
    
    print("✓ _Settings sheet created")
    
    # ============================================================================
    # FRONTEND SHEETS (Visible)
    # ============================================================================
    
    # 1. Dashboard_Admin Sheet
    ws_dashboard = wb.create_sheet("Dashboard_Admin", 0)
    
    row = create_luxury_header(ws_dashboard, "ADMIN DASHBOARD", "Executive Overview & Business Intelligence")
    
    # KPIs Section
    ws_dashboard[f'A{row}'].value = "KEY PERFORMANCE INDICATORS"
    ws_dashboard[f'A{row}'].font = FONTS['subheader']
    ws_dashboard[f'A{row}'].fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    ws_dashboard.merge_cells(f'A{row}:D{row}')
    row += 1
    
    kpi_data = [
        ["Metric", "Value", "Target", "Status"],
        ["Today's Revenue", "AED 45,000", "AED 40,000", "✓ On Track"],
        ["Monthly Revenue", "AED 850,000", "AED 800,000", "✓ Exceeding"],
        ["Inventory Value", "AED 500,000", "-", "Available"],
        ["Active Rentals", "12", "-", "Normal"],
        ["VIP Customers", "45", "-", "Growing"]
    ]
    
    for data_row in kpi_data:
        ws_dashboard.append(data_row)
    
    apply_table_style(ws_dashboard, row, row + len(kpi_data) - 2, 1, 4)
    
    row += len(kpi_data) + 2
    
    # Quick Actions Section
    ws_dashboard[f'A{row}'].value = "QUICK ACTIONS"
    ws_dashboard[f'A{row}'].font = FONTS['subheader']
    ws_dashboard[f'A{row}'].fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    ws_dashboard.merge_cells(f'A{row}:D{row}')
    row += 2
    
    quick_actions = [
        "[New Sale]",
        "[Add Product]",
        "[View Reports]",
        "[Stock Adjustment]",
        "[Customer Management]",
        "[User Settings]"
    ]
    
    for action in quick_actions:
        ws_dashboard[f'A{row}'].value = action
        ws_dashboard[f'A{row}'].font = Font(name='Calibri', size=10, color='0070C0', underline='single')
        row += 1
    
    # Hide gridlines
    ws_dashboard.sheet_view.showGridLines = False
    
    print("✓ Dashboard_Admin sheet created")
    
    # 2. POS Sheet (Sales Module)
    ws_pos = wb.create_sheet("POS")
    
    row = create_luxury_header(ws_pos, "POINT OF SALE", "Fast & Intuitive Sales Interface")
    
    ws_pos[f'A{row}'].value = "CUSTOMER"
    ws_pos[f'A{row}'].font = FONTS['label']
    ws_pos[f'B{row}'].value = "[Select or Create New]"
    row += 2
    
    ws_pos[f'A{row}'].value = "ITEMS"
    ws_pos[f'A{row}'].font = FONTS['label']
    row += 1
    
    # Items table
    item_headers = ["Product", "Qty", "Price (AED)", "Type", "Total (AED)", ""]
    for col, header in enumerate(item_headers, 1):
        cell = ws_pos.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    row += 1
    
    # Store row references for formulas
    items_start_row = row
    for i in range(10):
        ws_pos.cell(row=row+i, column=1).value = ""
        ws_pos.cell(row=row+i, column=6).value = "[Delete]"
    
    items_end_row = row + 9
    row += 12
    
    # Summary Section
    ws_pos[f'A{row}'].value = "SUBTOTAL"
    subtotal_row = row
    ws_pos[f'B{row}'].value = f"=SUM(E{items_start_row}:E{items_end_row})"
    ws_pos[f'B{row}'].number_format = 'AED #,##0.00'
    row += 1
    
    ws_pos[f'A{row}'].value = "VAT (5%)"
    vat_row = row
    ws_pos[f'B{row}'].value = f"=B{subtotal_row}*0.05"
    ws_pos[f'B{row}'].number_format = 'AED #,##0.00'
    row += 1
    
    ws_pos[f'A{row}'].value = "DISCOUNT"
    discount_row = row
    ws_pos[f'B{row}'].value = 0
    ws_pos[f'B{row}'].number_format = 'AED #,##0.00'
    row += 1
    
    ws_pos[f'A{row}'].value = "TOTAL"
    ws_pos[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=COLORS['gold'])
    ws_pos[f'B{row}'].value = f"=B{subtotal_row}+B{vat_row}-B{discount_row}"
    ws_pos[f'B{row}'].font = Font(name='Calibri', size=12, bold=True)
    ws_pos[f'B{row}'].number_format = 'AED #,##0.00'
    row += 2
    
    ws_pos[f'A{row}'].value = "PAYMENT METHOD"
    ws_pos[f'B{row}'].value = "[Select Method]"
    row += 2
    
    ws_pos[f'A{row}'].value = "[Complete Payment]"
    ws_pos[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws_pos[f'A{row}'].fill = PatternFill(start_color=COLORS['charcoal'], end_color=COLORS['charcoal'], fill_type='solid')
    ws_pos.merge_cells(f'A{row}:B{row}')
    
    ws_pos.sheet_view.showGridLines = False
    
    print("✓ POS sheet created")
    
    # 3. Inventory Sheet
    ws_inventory = wb.create_sheet("Inventory")
    
    row = create_luxury_header(ws_inventory, "INVENTORY MANAGEMENT", "Stock Levels & Movement Control")
    
    ws_inventory[f'A{row}'].value = "Current Inventory"
    ws_inventory[f'A{row}'].font = FONTS['subheader']
    ws_inventory.merge_cells(f'A{row}:E{row}')
    row += 1
    
    inv_headers = ["Product ID", "Product Name", "Current Stock", "Low Stock Level", "Last Updated"]
    for col, header in enumerate(inv_headers, 1):
        cell = ws_inventory.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    row += 1
    apply_table_style(ws_inventory, row, row + 19, 1, 5)
    
    for i in range(20):
        ws_inventory.cell(row=row+i, column=1).value = ""
    
    row += 22
    
    ws_inventory[f'A{row}'].value = "Stock Movement Log"
    ws_inventory[f'A{row}'].font = FONTS['subheader']
    ws_inventory.merge_cells(f'A{row}:F{row}')
    row += 1
    
    mov_headers = ["Movement ID", "Product", "Type", "Qty", "Date", "User"]
    for col, header in enumerate(mov_headers, 1):
        cell = ws_inventory.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    ws_inventory.sheet_view.showGridLines = False
    
    print("✓ Inventory sheet created")
    
    # 4. Products Sheet
    ws_products_ui = wb.create_sheet("Products")
    
    row = create_luxury_header(ws_products_ui, "PRODUCT CATALOG", "Manage Dresses, Images & Barcodes")
    
    prod_headers = ["Product ID", "Name", "Category", "Size", "Color", "Fabric", "Cost Price", "Selling Price"]
    for col, header in enumerate(prod_headers, 1):
        cell = ws_products_ui.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    row += 1
    apply_table_style(ws_products_ui, row, row + 19, 1, 8)
    
    row += 22
    
    ws_products_ui[f'A{row}'].value = "[+ Add New Product]"
    ws_products_ui[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    ws_products_ui[f'A{row}'].fill = PatternFill(start_color=COLORS['charcoal'], end_color=COLORS['charcoal'], fill_type='solid')
    ws_products_ui.merge_cells(f'A{row}:B{row}')
    
    ws_products_ui.sheet_view.showGridLines = False
    
    print("✓ Products sheet created")
    
    # 5. Customers Sheet
    ws_cust = wb.create_sheet("Customers")
    
    row = create_luxury_header(ws_cust, "CUSTOMER RELATIONSHIP MANAGEMENT", "Track Clients, Preferences & History")
    
    cust_headers = ["Customer ID", "Name", "Phone", "Email", "Total Spending", "Visit Count", "VIP Status", "Last Visit"]
    for col, header in enumerate(cust_headers, 1):
        cell = ws_cust.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    row += 1
    apply_table_style(ws_cust, row, row + 19, 1, 8)
    
    row += 22
    
    ws_cust[f'A{row}'].value = "[+ Add New Customer]"
    ws_cust[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    ws_cust[f'A{row}'].fill = PatternFill(start_color=COLORS['charcoal'], end_color=COLORS['charcoal'], fill_type='solid')
    ws_cust.merge_cells(f'A{row}:B{row}')
    
    ws_cust.sheet_view.showGridLines = False
    
    print("✓ Customers sheet created")
    
    # 6. Reports Sheet
    ws_reports = wb.create_sheet("Reports")
    
    row = create_luxury_header(ws_reports, "BUSINESS REPORTS", "Analytics, Insights & Trends")
    
    ws_reports[f'A{row}'].value = "Daily Sales Summary"
    ws_reports[f'A{row}'].font = FONTS['subheader']
    row += 1
    
    ws_reports[f'A{row}'].value = "Date: " + datetime.datetime.now().strftime("%d/%m/%Y")
    row += 1
    
    report_headers = ["Time", "Transaction ID", "Customer", "Amount", "Payment", "Items"]
    for col, header in enumerate(report_headers, 1):
        cell = ws_reports.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    row += 1
    apply_table_style(ws_reports, row, row + 9, 1, 6)
    
    row += 12
    
    ws_reports[f'A{row}'].value = "TOTAL REVENUE (Today)"
    ws_reports[f'B{row}'].value = "AED 0"
    ws_reports[f'B{row}'].font = Font(name='Calibri', size=11, bold=True)
    
    ws_reports.sheet_view.showGridLines = False
    
    print("✓ Reports sheet created")
    
    # 7. Financial_Ledger Sheet
    ws_ledger = wb.create_sheet("Financial_Ledger")
    
    row = create_luxury_header(ws_ledger, "ACCOUNTING LEDGER", "Financial Transactions & Balance Sheet")
    
    ledger_headers = ["Date", "Reference", "Account Type", "Debit (AED)", "Credit (AED)", "Balance (AED)"]
    for col, header in enumerate(ledger_headers, 1):
        cell = ws_ledger.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    row += 1
    ledger_start_row = row
    ledger_end_row = row + 19
    
    # Add formulas for Balance column (running total)
    for i in range(20):
        current_row = row + i
        ws_ledger.cell(row=current_row, column=4).number_format = 'AED #,##0.00'
        ws_ledger.cell(row=current_row, column=5).number_format = 'AED #,##0.00'
        
        # Balance = previous balance + debit - credit
        if i == 0:
            ws_ledger.cell(row=current_row, column=6).value = f"=D{current_row}-E{current_row}"
        else:
            ws_ledger.cell(row=current_row, column=6).value = f"=F{current_row-1}+D{current_row}-E{current_row}"
        ws_ledger.cell(row=current_row, column=6).number_format = 'AED #,##0.00'
    
    apply_table_style(ws_ledger, ledger_start_row, ledger_end_row, 1, 6)
    
    # Add date picker to Date column
    ledger_date_validation = DataValidation(type="date", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
    ledger_date_validation.error = 'Please enter a valid date'
    ledger_date_validation.errorTitle = 'Invalid Date'
    ws_ledger.add_data_validation(ledger_date_validation)
    ledger_date_validation.add(f'A{ledger_start_row}:A{ledger_end_row}')
    
    ws_ledger.sheet_view.showGridLines = False
    
    print("✓ Financial_Ledger sheet created with formulas and date picker")
    
    # 8. Cash_Flow Sheet
    ws_cashflow = wb.create_sheet("Cash_Flow")
    
    row = create_luxury_header(ws_cashflow, "CASH FLOW ANALYSIS", "Daily, Weekly & Monthly Tracking")
    
    cf_headers = ["Date", "Description", "Cash In (AED)", "Cash Out (AED)", "Net (AED)", "Balance (AED)"]
    for col, header in enumerate(cf_headers, 1):
        cell = ws_cashflow.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    row += 1
    cashflow_start_row = row
    cashflow_end_row = row + 14
    
    # Add formulas for Net and Balance columns
    for i in range(15):
        current_row = row + i
        # Net = Cash In - Cash Out
        ws_cashflow.cell(row=current_row, column=5).value = f"=C{current_row}-D{current_row}"
        ws_cashflow.cell(row=current_row, column=5).number_format = 'AED #,##0.00'
        
        # Balance (running total)
        if i == 0:
            ws_cashflow.cell(row=current_row, column=6).value = f"=E{current_row}"
        else:
            ws_cashflow.cell(row=current_row, column=6).value = f"=E{current_row}+F{current_row-1}"
        ws_cashflow.cell(row=current_row, column=6).number_format = 'AED #,##0.00'
        
        # Add currency format to Cash In and Cash Out
        ws_cashflow.cell(row=current_row, column=3).number_format = 'AED #,##0.00'
        ws_cashflow.cell(row=current_row, column=4).number_format = 'AED #,##0.00'
    
    apply_table_style(ws_cashflow, cashflow_start_row, cashflow_end_row, 1, 6)
    
    # Add date picker to Date column
    date_validation = DataValidation(type="date", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
    date_validation.error = 'Please enter a valid date'
    date_validation.errorTitle = 'Invalid Date'
    ws_cashflow.add_data_validation(date_validation)
    date_validation.add(f'A{cashflow_start_row}:A{cashflow_end_row}')
    
    ws_cashflow.sheet_view.showGridLines = False
    
    print("✓ Cash_Flow sheet created with formulas and date picker")
    
    # 9. Dashboard_Sales Sheet
    ws_dash_sales = wb.create_sheet("Dashboard_Sales")
    
    row = create_luxury_header(ws_dash_sales, "SALES DASHBOARD", "Personal Sales Performance & Targets")
    
    ws_dash_sales[f'A{row}'].value = "Your Performance Today"
    ws_dash_sales[f'A{row}'].font = FONTS['subheader']
    row += 1
    
    perf_data = [
        ["Metric", "Value"],
        ["Sales Count", "8"],
        ["Total Revenue", "AED 125,000"],
        ["Average Transaction", "AED 15,625"],
        ["Rentals Booked", "3"]
    ]
    
    for data in perf_data:
        ws_dash_sales.append(data)
    
    ws_dash_sales.sheet_view.showGridLines = False
    
    print("✓ Dashboard_Sales sheet created")
    
    # 10. Stock_Movements Sheet
    ws_stock_mov = wb.create_sheet("Stock_Movements")
    
    row = create_luxury_header(ws_stock_mov, "STOCK MOVEMENTS LOG", "Complete History of Inventory Changes")
    
    mov_headers = ["Movement ID", "Product ID", "Product Name", "Type", "Quantity", "Reference", "Date", "User"]
    for col, header in enumerate(mov_headers, 1):
        cell = ws_stock_mov.cell(row=row, column=col)
        cell.value = header
        cell.font = FONTS['subheader']
        cell.fill = PatternFill(start_color=COLORS['gold'], end_color=COLORS['gold'], fill_type='solid')
    
    row += 1
    mov_start_row = row
    mov_end_row = row + 19
    
    apply_table_style(ws_stock_mov, mov_start_row, mov_end_row, 1, 8)
    
    # Add date picker to Date column (column G)
    mov_date_validation = DataValidation(type="date", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
    mov_date_validation.error = 'Please enter a valid date'
    mov_date_validation.errorTitle = 'Invalid Date'
    ws_stock_mov.add_data_validation(mov_date_validation)
    mov_date_validation.add(f'G{mov_start_row}:G{mov_end_row}')
    
    ws_stock_mov.sheet_view.showGridLines = False
    
    print("✓ Stock_Movements sheet created with date picker")
    
    # Set column widths for all sheets
    for ws in wb.sheetnames:
        sheet = wb[ws]
        for column in sheet.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def save_workbook(wb, filepath):
    """Save workbook to file"""
    wb.save(filepath)
    print(f"\n✓ Workbook saved to: {filepath}")

if __name__ == "__main__":
    print("\n" + "="*60)
    print("RIAMAN FASHION ERP WORKBOOK GENERATOR")
    print("="*60 + "\n")
    
    wb = create_workbook()
    filepath = r"C:\Users\KAIS\Documents\RIAMAN_FASHION_ERP\RIAMAN_FASHION_ERP.xlsx"
    save_workbook(wb, filepath)
    
    print("\n" + "="*60)
    print("✓ WORKBOOK CREATION COMPLETE!")
    print("="*60)
    print("\nNext steps:")
    print("1. Import VBA modules from RIAMAN_FASHION_ERP_Main.vbs")
    print("2. Add product images to Products sheet")
    print("3. Configure login form")
    print("4. Test role-based access")
    print("5. Protect all backend sheets")
    print("="*60 + "\n")
