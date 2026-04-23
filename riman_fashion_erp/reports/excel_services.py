"""
Excel export and import services
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.conf import settings
from datetime import datetime
from inventory.models import Product
from crm.models import Client
from django.db import transaction


class ExcelExportService:
    """Generate Excel exports for various reports"""
    
    HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def export_sales_report(sales, date_from=None, date_to=None):
        """Export sales report to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Headers
        headers = ['Sale #', 'Customer', 'Email', 'Phone', 'Sale Date', 'Items', 'Subtotal', 'Tax', 'Total', 'Status']
        ws.append(headers)
        
        # Format headers
        for cell in ws[1]:
            cell.fill = ExcelExportService.HEADER_FILL
            cell.font = ExcelExportService.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = ExcelExportService.BORDER
        
        # Data rows
        for sale in sales:
            ws.append([
                sale.sale_number,
                f"{sale.customer.first_name} {sale.customer.last_name}",
                getattr(sale.customer, 'email', ''),
                getattr(sale.customer, 'phone', ''),
                sale.sale_date.strftime('%m/%d/%Y'),
                len(sale.lines.all()),
                f"{sale.subtotal:.2f}",
                f"{sale.tax_amount:.2f}",
                f"{sale.total_amount:.2f}",
                sale.payment_status,
            ])
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = ExcelExportService.BORDER
                cell.alignment = Alignment(horizontal='left')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(['Metric', 'Value'])
        ws_summary.append(['Total Sales', len(sales)])
        ws_summary.append(['Total Revenue', f"{sum(s.total_amount for s in sales):.2f}"])
        ws_summary.append(['Total Items', sum(len(s.lines.all()) for s in sales)])
        ws_summary.append(['Date From', date_from.strftime('%m/%d/%Y') if date_from else 'All'])
        ws_summary.append(['Date To', date_to.strftime('%m/%d/%Y') if date_to else 'All'])
        
        for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row):
            for cell in row:
                cell.border = ExcelExportService.BORDER
        
        ws_summary.column_dimensions['A'].width = 15
        ws_summary.column_dimensions['B'].width = 20
        
        # Return as BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_invoices(invoices):
        """Export invoices to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        
        # Headers
        headers = ['Invoice #', 'Sale #', 'Customer', 'Invoice Date', 'Total', 'Status']
        ws.append(headers)
        
        # Format headers
        for cell in ws[1]:
            cell.fill = ExcelExportService.HEADER_FILL
            cell.font = ExcelExportService.HEADER_FONT
            cell.border = ExcelExportService.BORDER
        
        # Data
        for invoice in invoices:
            ws.append([
                invoice.invoice_number,
                invoice.sale.sale_number,
                f"{invoice.sale.customer.first_name} {invoice.sale.customer.last_name}",
                invoice.invoice_date.strftime('%m/%d/%Y'),
                f"{invoice.sale.total_amount:.2f}",
                invoice.status,
            ])
        
        # Adjust columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_gl_report(journal_entries):
        """Export GL to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "General Ledger"
        
        # Headers
        headers = ['Entry #', 'Date', 'Description', 'Type', 'Account', 'Debit', 'Credit', 'Status', 'Balanced']
        ws.append(headers)
        
        # Format headers
        for cell in ws[1]:
            cell.fill = ExcelExportService.HEADER_FILL
            cell.font = ExcelExportService.HEADER_FONT
            cell.border = ExcelExportService.BORDER
        
        # Data
        row_num = 2
        for entry in journal_entries:
            is_balanced = entry.is_balanced()
            
            for line in entry.lines.all():
                ws.append([
                    entry.journal_number if line == entry.lines.first() else '',
                    entry.entry_date.strftime('%m/%d/%Y') if line == entry.lines.first() else '',
                    entry.description if line == entry.lines.first() else '',
                    entry.entry_type if line == entry.lines.first() else '',
                    line.account.account_name,
                    f"{line.amount:.2f}" if line.line_type == 'debit' else '',
                    f"{line.amount:.2f}" if line.line_type == 'credit' else '',
                    entry.status if line == entry.lines.first() else '',
                    'Yes' if is_balanced and line == entry.lines.first() else ('No' if line == entry.lines.first() else ''),
                ])
                row_num += 1
        
        # Adjust columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_inventory(products):
        """Export product inventory to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        # Headers
        headers = ['SKU', 'Product Name', 'Category', 'Quantity In Stock', 'Cost Price', 'Sale Price', 'Rental Price', 'Status']
        ws.append(headers)
        
        # Format headers
        for cell in ws[1]:
            cell.fill = ExcelExportService.HEADER_FILL
            cell.font = ExcelExportService.HEADER_FONT
            cell.border = ExcelExportService.BORDER
        
        # Data
        for product in products:
            ws.append([
                product.sku,
                product.name,
                str(product.category),
                product.quantity_in_stock,
                f"{product.cost_price:.2f}",
                f"{product.sale_price:.2f}",
                f"{product.rental_price:.2f}" if hasattr(product, 'rental_price') else '',
                'Active' if product.is_active else 'Inactive',
            ])
        
        # Adjust columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


class ExcelImportService:
    """Import data from Excel files"""
    
    @staticmethod
    def import_products(file_path):
        """Import products from Excel"""
        wb = load_workbook(file_path)
        ws = wb.active
        
        imported = 0
        errors = []
        
        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Skip empty rows
                    if not row or all(cell is None for cell in row):
                        continue
                    
                    # Unpack only the first 8 columns we need
                    sku, name, category, qty, cost, sale_price, rental_price, is_active = row[:8]
                    
                    product, created = Product.objects.update_or_create(
                        sku=sku,
                        defaults={
                            'name': name,
                            'category': category,
                            'quantity_in_stock': int(qty),
                            'cost_price': float(cost),
                            'sale_price': float(sale_price),
                            'is_active': is_active.lower() in ['true', 'yes', '1'],
                        }
                    )
                    
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
        
        return {'imported': imported, 'errors': errors}
    
    @staticmethod
    def import_customers(file_path):
        """Import customers from Excel"""
        wb = load_workbook(file_path)
        ws = wb.active
        
        imported = 0
        errors = []
        
        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Skip empty rows
                    if not row or all(cell is None for cell in row):
                        continue
                    
                    # Unpack only the first 8 columns we need
                    first_name, last_name, email, phone, address, city, state, country = row[:8]
                    
                    client, created = Client.objects.update_or_create(
                        email=email,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'phone': phone,
                            'address': address,
                            'city': city,
                            'state': state,
                            'country': country,
                        }
                    )
                    
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
        
        return {'imported': imported, 'errors': errors}
